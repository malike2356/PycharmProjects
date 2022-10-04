# Find more on openpyxl in the link below:
# https://openpyxl.readthedocs.io/en/stable/tutorial.html


import openpyxl as xl
from openpyxl.chart import BarChart, PieChart, LineChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]  # Putting the sheet in a list
    total_rows = sheet.max_row  # find the total number rows and assign to variable

    for row in range(2, total_rows + 1):  # plus 1 to include the row number itself since index starts from 0
        cell = sheet.cell(row, 3)  # sheet.cell(x, y) # x is row, y is column, OR cell = sheet["a1"] same result
        corrected_price = cell.value * 0.9  # 9 percent correction
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    chart_values = Reference(sheet,
                             min_row=2,
                             max_row=total_rows,
                             min_col=4,
                             max_col=4)
    chart = BarChart()
    chart.add_data(chart_values)
    sheet.add_chart(chart, "e2")
    wb.save(filename)


